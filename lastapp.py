import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import cv2, numpy as np, math, os, csv, json, datetime
from PIL import Image, ImageTk

def clamp(v, lo, hi): return max(lo, min(hi, v))
def hsv_to_bgr(h, s, v):
    hsv = np.uint8([[[int(h)%180, clamp(int(s),0,255), clamp(int(v),0,255)]]])
    bgr = cv2.cvtColor(hsv, cv2.COLOR_HSV2BGR)[0][0]
    return (int(bgr[0]), int(bgr[1]), int(bgr[2]))
def mean_hue(ranges):
    best_span=-1; best_mid=0
    for lo,hi in ranges:
        h1,h2=int(lo[0]), int(hi[0])
        span=(h2-h1)%180
        mid=(h1+span//2)%180
        if span>best_span: best_span, best_mid = span, mid
    return best_mid
def build_palette(base_h):
    comp=(base_h+90)%180; s,v=220,230
    return {
        "area_fill":    hsv_to_bgr(comp, int(s*0.7), int(v*0.8)),
        "area_edge":    hsv_to_bgr(comp, s, v),
        "poly_edge":    hsv_to_bgr((comp+18)%180, s, v),
        "poly_vert":    hsv_to_bgr((comp-18)%180, s, v),
        "inner_circle": hsv_to_bgr((comp+36)%180, s, v),
        "ghost_edge":   (255,255,0),
        "ghost_poly":   (0,255,255),
        "ok":           (0,255,0),
        "bad":          (0,0,255),
        "text":         (240,240,240),
    }
def inner_tangent_circle(mask):
    dt = cv2.distanceTransform(mask, cv2.DIST_L2, 5)
    _, mx, _, _ = cv2.minMaxLoc(dt)
    if mx <= 0: return [], 0, dt
    H,W = dt.shape
    tol = max(1.0, mx*0.002)
    cs=[]
    for y in range(H):
        for x in range(W):
            if abs(dt[y,x]-mx) <= tol: cs.append((x,y))
    return cs, int(round(mx)), dt
def poly_area(pts):
    s=0.0; n=len(pts)
    for i in range(n):
        x1,y1=pts[i]; x2,y2=pts[(i+1)%n]
        s += x1*y2 - x2*y1
    return abs(s)/2.0
def poly_perimeter(pts):
    p=0.0; n=len(pts)
    for i in range(n):
        x1,y1=pts[i]; x2,y2=pts[(i+1)%n]
        p += math.hypot(x2-x1, y2-y1)
    return p
def polygon_mask(vertices, hw):
    H,W = hw
    m = np.zeros((H,W), np.uint8)
    if len(vertices)>=3:
        cv2.fillPoly(m, [np.array(vertices, np.int32)], 255)
    return m
def point_in_poly(pt, vertices):
    cnt = np.array(vertices, np.int32)
    return cv2.pointPolygonTest(cnt, (float(pt[0]), float(pt[1])), False) >= 0
def centroid(vertices):
    A=0.0; Cx=0.0; Cy=0.0
    for i in range(len(vertices)):
        x1,y1=vertices[i]; x2,y2=vertices[(i+1)%len(vertices)]
        c=x1*y2 - x2*y1
        A+=c; Cx+=(x1+x2)*c; Cy+=(y1+y2)*c
    if abs(A)<1e-6:
        xs=[p[0] for p in vertices]; ys=[p[1] for p in vertices]
        return (sum(xs)/len(xs), sum(ys)/len(ys))
    A*=0.5; Cx/=(6*A); Cy/=(6*A)
    return (Cx, Cy)
def rotate_points(vertices, center, angle_rad):
    cx,cy=center; co=math.cos(angle_rad); si=math.sin(angle_rad)
    out=[]
    for x,y in vertices:
        dx,dy=x-cx, y-cy
        out.append((cx+dx*co-dy*si, cy+dx*si+dy*co))
    return out
def eroded_overlap(maskA, maskB, iters=1):
    k = np.ones((3,3), np.uint8)
    erA = cv2.erode(maskA, k, iterations=iters)
    erB = cv2.erode(maskB, k, iterations=iters)
    inter = cv2.bitwise_and(erA, erB)
    return cv2.countNonZero(inter) > 0
def p2seg_dist(p,a,b):
    px,py=p; ax,ay=a; bx,by=b
    vx=bx-ax; vy=by-ay
    wx=px-ax; wy=py-ay
    vv=vx*vx+vy*vy
    if vv==0: return math.hypot(px-ax, py-ay)
    t=(wx*vx+wy*vy)/vv; t=clamp(t,0.0,1.0)
    proj=(ax+t*vx, ay+t*vy)
    return math.hypot(px-proj[0], py-proj[1])
def min_contour_distance(A_pts, B_pts):
    if len(A_pts)==0 or len(B_pts)==0: return float('inf')
    A=A_pts.reshape(-1,2); B=B_pts.reshape(-1,2)
    md=float('inf')
    for p in A:
        for j in range(len(B)):
            md=min(md, p2seg_dist((float(p[0]),float(p[1])),
                                  (float(B[j,0]),float(B[j,1])),
                                  (float(B[(j+1)%len(B),0]),float(B[(j+1)%len(B),1]))))
    for p in B:
        for j in range(len(A)):
            md=min(md, p2seg_dist((float(p[0]),float(p[1])),
                                  (float(A[j,0]),float(A[j,1])),
                                  (float(A[(j+1)%len(A),0]),float(A[(j+1)%len(A),1]))))
    return md
def signed_area(vertices):
    s=0.0
    for i in range(len(vertices)):
        x1,y1=vertices[i]; x2,y2=vertices[(i+1)%len(vertices)]
        s += x1*y2 - x2*y1
    return 0.5*s
def simplify_chain(pts, eps=1.0):
    out=[]
    for p in pts:
        if not out: out.append(p); continue
        if math.hypot(p[0]-out[-1][0], p[1]-out[-1][1]) >= eps: out.append(p)
    if len(out)>=2 and math.hypot(out[0][0]-out[-1][0], out[0][1]-out[-1][1])<eps:
        out[-1]=out[0]
    return out

COLORS = {
    "Red":[(np.array([0,70,50],np.uint8),   np.array([10,255,255],np.uint8)),
           (np.array([170,70,50],np.uint8), np.array([180,255,255],np.uint8))],
    "Orange":[(np.array([10,100,40],np.uint8), np.array([25,255,255],np.uint8))],
    "Yellow":[(np.array([25,70,70],np.uint8),  np.array([35,255,255],np.uint8))],
    "Green":[(np.array([35,50,50],np.uint8),   np.array([85,255,255],np.uint8))],
    "Blue":[(np.array([100,60,40],np.uint8),   np.array([140,255,255],np.uint8))],
    "Purple":[(np.array([140,50,50],np.uint8), np.array([160,255,255],np.uint8))],
    "Pink":[(np.array([150,40,80],np.uint8),   np.array([170,255,255],np.uint8))],
    "White":[(np.array([0,0,200],np.uint8),    np.array([180,40,255],np.uint8))],
    "Black":[(np.array([0,0,0],np.uint8),      np.array([180,255,50],np.uint8))],
    "Gray":[(np.array([0,0,50],np.uint8),      np.array([180,40,200],np.uint8))],
    "Brown":[(np.array([10,120,20],np.uint8),  np.array([25,255,140],np.uint8))],
}

class Area:
    def __init__(self, area_id, name, mask, contour, base_h):
        self.id=area_id; self.name=name
        self.mask=mask
        self.dt=cv2.distanceTransform(mask, cv2.DIST_L2, 5)
        self.contour=contour
        self.base_h=base_h
        self.center=None; self.rmax=None
        self.polygon=None
        self.polygon_mask=None
        self.included=True; self.exclude_reason=None
        self.proximity_px=None; self.proximity_cm=None

class SmartPolygonApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Smart Polygon App")
        self.geometry("1320x900")
        self.configure(bg="#f0f0f0")
        self.step=0; self.input_type=None; self.path=None
        self.image=None
        self.full_color_mask=None
        self.mask=None
        self.dt_global=None
        self.contour=None
        self.main_mask=None
        self.main_contour=None
        self.px_to_cm=None; self.measure_val=None
        self.choice=None
        self.draw_pts=[]; self.cursor_pt=None
        self.curr_area_mask=None; self.curr_contour=None; self.curr_dt=None
        self.centers=[]; self.sel_center=None; self.radius=None
        self.poly_method=None; self._poly_clicks=[]
        self.areas=[]; self.area_counter=0
        self.wizard_on=tk.BooleanVar(value=False)
        self.enforce_local=tk.BooleanVar(value=True)
        self.enforce_global=tk.BooleanVar(value=True)
        self.rotate_step=tk.DoubleVar(value=1.0)
        self.exclude_if_far=tk.BooleanVar(value=True)
        self.proximity_px=tk.IntVar(value=15)
        self.prox_ref=tk.StringVar(value="Main Color Area")
        self.contrast=tk.BooleanVar(value=True)
        self.inside_margin_px=tk.DoubleVar(value=1.0)
        self.selected_area=None
        self.dragging=False; self.drag_start=None
        self.wiz_orig=None; self.wiz_pivot=None; self.wiz_dx=0; self.wiz_dy=0; self.wiz_angle=0.0
        self.snap_active=False
        self.drag_preview=None; self.drag_valid=False
        self.snap_px=10
        self.zoom=1.0; self.min_zoom=0.25; self.max_zoom=6.0
        self.merge_state=None
        self.merge_base=None
        self.merge_move=None
        self.merge_init_vertices=None
        self.merge_best=(None,None)
        self.merge_dx=0.0; self.merge_dy=0.0; self.merge_angle=0.0
        self.merge_pivot=(0.0,0.0)
        self.merge_tol_angle=tk.DoubleVar(value=1.5)
        self.merge_tol_offset=tk.DoubleVar(value=3.0)
        self.merge_tol_endgap=tk.DoubleVar(value=3.0)
        self._build_ui()

    def _build_ui(self):
        left=ttk.Frame(self); left.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        self.step_lbl=ttk.Label(left, text="Step 0"); self.step_lbl.pack(pady=3)
        self.info_lbl=ttk.Label(left, text="Welcome! Click Next.", wraplength=260); self.info_lbl.pack(pady=3)
        ttk.Checkbutton(left, text="Wizard (Move/Rotate)", variable=self.wizard_on, command=self._wizard_toggle).pack(pady=(6,3))
        wbox=ttk.LabelFrame(left, text="Wizard Options"); wbox.pack(fill=tk.X, pady=(0,6))
        ttk.Checkbutton(wbox, text="Stay inside local area", variable=self.enforce_local, command=self._redraw).pack(anchor="w", padx=6, pady=2)
        ttk.Checkbutton(wbox, text="Stay inside global mask", variable=self.enforce_global, command=self._redraw).pack(anchor="w", padx=6, pady=2)
        rrow=ttk.Frame(wbox); rrow.pack(fill=tk.X, pady=3)
        ttk.Label(rrow, text="Rotate step (°):").pack(side=tk.LEFT, padx=6)
        ttk.Spinbox(rrow, from_=0.1, to=15.0, increment=0.1, textvariable=self.rotate_step, width=6).pack(side=tk.LEFT)
        mbox=ttk.LabelFrame(left, text="Merge"); mbox.pack(fill=tk.X, pady=(0,6))
        self.btn_prepare=ttk.Button(mbox, text="Prepare Merge…", command=self._merge_prepare_prompt)
        self.btn_prepare.pack(fill=tk.X, padx=6, pady=(4,2))
        self.btn_commit=ttk.Button(mbox, text="Commit Merge", command=self._merge_commit, state="disabled")
        self.btn_commit.pack(fill=tk.X, padx=6, pady=2)
        self.btn_cancel=ttk.Button(mbox, text="Cancel Align", command=self._merge_cancel, state="disabled")
        self.btn_cancel.pack(fill=tk.X, padx=6, pady=(2,6))
        trow=ttk.Frame(mbox); trow.pack(fill=tk.X, pady=(0,4))
        ttk.Label(trow, text="Tol angle (°):").pack(side=tk.LEFT, padx=6)
        ttk.Spinbox(trow, from_=0.1, to=5.0, increment=0.1, textvariable=self.merge_tol_angle, width=6).pack(side=tk.LEFT)
        trow2=ttk.Frame(mbox); trow2.pack(fill=tk.X, pady=(0,4))
        ttk.Label(trow2, text="Tol offset (px):").pack(side=tk.LEFT, padx=6)
        ttk.Spinbox(trow2, from_=0.5, to=10.0, increment=0.5, textvariable=self.merge_tol_offset, width=6).pack(side=tk.LEFT)
        trow3=ttk.Frame(mbox); trow3.pack(fill=tk.X, pady=(0,4))
        ttk.Label(trow3, text="Tol endgap (px):").pack(side=tk.LEFT, padx=6)
        ttk.Spinbox(trow3, from_=0.5, to=10.0, increment=0.5, textvariable=self.merge_tol_endgap, width=6).pack(side=tk.LEFT)
        pbox=ttk.LabelFrame(left, text="Proximity Filter"); pbox.pack(fill=tk.X, pady=(0,6))
        ttk.Checkbutton(pbox, text="Exclude if far", variable=self.exclude_if_far, command=self._recalc_inclusions).pack(anchor="w", padx=6, pady=(4,2))
        rr2=ttk.Frame(pbox); rr2.pack(fill=tk.X)
        ttk.Label(rr2, text="Threshold (px):").pack(side=tk.LEFT, padx=6)
        ttk.Spinbox(rr2, from_=1, to=300, increment=1, textvariable=self.proximity_px, width=6, command=self._recalc_inclusions).pack(side=tk.LEFT)
        rr3=ttk.Frame(pbox); rr3.pack(fill=tk.X, pady=3)
        ttk.Label(rr3, text="Reference:").pack(side=tk.LEFT, padx=6)
        ttk.OptionMenu(rr3, self.prox_ref, "Main Color Area", "Main Color Area", "Local Area", command=lambda *_: self._recalc_inclusions()).pack(side=tk.LEFT, padx=4)
        ttk.Checkbutton(left, text="Contrast preview", variable=self.contrast, command=self._redraw).pack(anchor="w", pady=(2,6))
        nav=ttk.Frame(left); nav.pack(side=tk.BOTTOM, fill=tk.X, pady=8)
        self.btn_back=ttk.Button(nav, text="Back", command=self._back); self.btn_back.pack(side=tk.LEFT, padx=4)
        self.btn_next=ttk.Button(nav, text="Next", command=self._next); self.btn_next.pack(side=tk.RIGHT, padx=4)
        right=ttk.Frame(self); right.pack(side=tk.RIGHT, fill=tk.BOTH, padx=5, pady=5)
        ttk.Label(right, text="Console:").pack(anchor="nw")
        self.log_txt=tk.Text(right, width=54, height=42); self.log_txt.pack(fill=tk.BOTH, expand=True)
        mid=ttk.Frame(self); mid.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.hbar=ttk.Scrollbar(mid, orient=tk.HORIZONTAL); self.hbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.vbar=ttk.Scrollbar(mid, orient=tk.VERTICAL); self.vbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas=tk.Canvas(mid, bg="black", xscrollcommand=self.hbar.set, yscrollcommand=self.vbar.set)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.hbar.config(command=self.canvas.xview); self.vbar.config(command=self.canvas.yview)
        self.canvas.bind("<ButtonPress-3>", self._pan_start)
        self.canvas.bind("<B3-Motion>", self._pan_drag)
        self.canvas.bind("<Control-MouseWheel>", self._zoom_wheel)
        self.canvas.bind("<Control-Button-4>", lambda e: self._zoom_generic(+1,e))
        self.canvas.bind("<Control-Button-5>", lambda e: self._zoom_generic(-1,e))

    def _log(self, msg):
        self.log_txt.insert(tk.END, msg+"\n"); self.log_txt.see(tk.END)

    def _next(self):
        if self.step==2 and self.measure_val is None:
            messagebox.showwarning("Warning","Please confirm the measurement first.")
            return
        if self.merge_state is not None:
            messagebox.showwarning("Warning","Finish or cancel merge alignment first.")
            return
        self.step+=1; self._run_step()

    def _back(self):
        if self.step>0: self.step-=1
        self._run_step()

    def _run_step(self):
        self.step_lbl.config(text=f"Step {self.step}")
        if self.step in (5,6,7,8):
            self.wizard_on.set(False); self._update_wizard_bindings()
        if self.step==0:
            self.info_lbl.config(text="Welcome! Click Next.")
            self.btn_back.config(state="disabled"); self.btn_next.config(state="normal")
            self.areas=[]; self.area_counter=0
            self.image=None; self.full_color_mask=None; self.mask=None; self.dt_global=None
            self.contour=None; self.main_mask=None; self.main_contour=None
            self.curr_area_mask=None; self.curr_contour=None; self.curr_dt=None
            self.draw_pts=[]; self.cursor_pt=None
            self.measure_val=None; self.px_to_cm=None
            self._redraw()
        elif self.step==1:
            self.btn_back.config(state="disabled")
            self.info_lbl.config(text="Choose a file (image or CSV).")
            self._choose_file()
        elif self.step==2:
            self.btn_back.config(state="normal")
            if self.input_type=="csv": self._measurement_csv()
            else: self._color_and_measurement()
        elif self.step==3:
            self._build_mask()
        elif self.step==4:
            self._area_choice()
        elif self.step==5:
            self._area_draw()
        elif self.step==6:
            self._inner_circle()
        elif self.step==7:
            self._polygon_method()
        elif self.step==8:
            self._polygon_create()
        elif self.step==9:
            self.info_lbl.config(text="Review & edit. Use Wizard and Prepare/Commit Merge. Click Next to export when finished.")
            self.wizard_on.set(True); self._update_wizard_bindings(); self._redraw()
        elif self.step==10:
            self.wizard_on.set(False); self._update_wizard_bindings()
            self._summary_and_export()
        else:
            self.info_lbl.config(text="Done.")
            self.btn_next.config(state="disabled"); self.btn_back.config(state="disabled")
        self._update_wizard_bindings()

    def _choose_file(self):
        self._log("Step 1: Selecting file…")
        path=filedialog.askopenfilename(
            filetypes=[("Supported","*.png;*.jpg;*.jpeg;*.bmp;*.csv"),
                       ("Image","*.png;*.jpg;*.jpeg;*.bmp"),
                       ("CSV","*.csv")])
        if not path:
            self._log("No file selected. Returning to Step 0.")
            self.step=0; self._run_step(); return
        self.path=path; ext=os.path.splitext(path)[1].lower()
        if ext==".csv":
            self.input_type="csv"; self._load_csv(path)
        else:
            self.input_type="image"
            img=cv2.imread(path)
            if img is None:
                messagebox.showerror("Error","Cannot open image."); self.step=0; self._run_step(); return
            self.image=img; self._log(f"Image loaded: {os.path.basename(path)}")
            self._redraw()

    def _load_csv(self, path):
        pts=[]
        with open(path,"r",encoding="utf-8",newline="") as f:
            sample=f.read(2048); f.seek(0)
            try: has_header=csv.Sniffer().has_header(sample)
            except: has_header=False
            try: dialect=csv.Sniffer().sniff(sample, delimiters=",; \t")
            except: dialect=csv.excel
            r=csv.reader(f, dialect)
            if has_header:
                head=next(r,[]); hl=[h.strip().lower() for h in head]
                ix=hl.index("x") if "x" in hl else 0
                iy=hl.index("y") if "y" in hl else 1
                for row in r:
                    if len(row)<=max(ix,iy): continue
                    try: pts.append((float(row[ix]), float(row[iy])))
                    except: pass
            else:
                for row in r:
                    if len(row)<2: continue
                    try: pts.append((float(row[0]), float(row[1])))
                    except: pass
        if len(pts)<3:
            messagebox.showerror("Error","CSV must contain at least 3 points.")
            self.step=0; self._run_step(); return
        xs=[p[0] for p in pts]; ys=[p[1] for p in pts]
        minx,miny=min(xs),min(ys); maxx,maxy=max(xs),max(ys)
        pad=40; W=int(math.ceil(maxx-minx))+pad; H=int(math.ceil(maxy-miny))+pad
        W=max(min(W,4000),300); H=max(min(H,4000),300)
        offx=-minx+pad/2.0; offy=-miny+pad/2.0
        pts_img=np.array([[int(round(x+offx)), int(round(y+offy))] for (x,y) in pts], np.int32)
        img=np.full((H,W,3),255,np.uint8)
        cv2.polylines(img,[pts_img],True,(0,0,255),2)
        m=np.zeros((H,W),np.uint8); cv2.fillPoly(m,[pts_img],255)
        self.image=img
        self.full_color_mask=m.copy()
        self.mask=m.copy()
        self.dt_global=cv2.distanceTransform(self.mask, cv2.DIST_L2, 5)
        self.contour=pts_img.reshape(-1,1,2)
        self.main_mask=self.mask.copy()
        self.main_contour=self.contour
        self._log(f"CSV polygon loaded: {len(pts)} points.")
        self._redraw()

    def _measurement_csv(self):
        self.info_lbl.config(text="Enter measurement value.")
        win=tk.Toplevel(self); win.title("Measurement"); win.geometry("300x170"); win.grab_set()
        tk.Label(win,text="Type:").pack()
        self.measure_type=tk.StringVar(value="Area")
        ttk.Radiobutton(win,text="Area (cm^2)",variable=self.measure_type,value="Area").pack(anchor="w")
        ttk.Radiobutton(win,text="Perimeter (cm)",variable=self.measure_type,value="Perimeter").pack(anchor="w")
        tk.Label(win,text="Value:").pack()
        ent=ttk.Entry(win); ent.insert(0,"100.0"); ent.pack()
        def ok():
            try:
                v=float(ent.get()); assert v>0
            except:
                messagebox.showerror("Error","Enter positive numeric value.",parent=win); return
            self.measure_val=v
            # >>> CSV modunda px_to_cm BURADA hesaplanır (kritik düzeltme) <<<
            if self.contour is None:
                messagebox.showerror("Error","CSV contour not available.",parent=win); return
            if self.measure_type.get()=="Area":
                A=cv2.contourArea(self.contour)
                if A<=0:
                    messagebox.showerror("Error","Zero area.",parent=win); return
                self.px_to_cm=math.sqrt(self.measure_val/A)
            else:
                P=cv2.arcLength(self.contour,True)
                if P<=0:
                    messagebox.showerror("Error","Zero perimeter.",parent=win); return
                self.px_to_cm=self.measure_val/P
            self._log(f"px->cm = {self.px_to_cm:.6f} (CSV)")
            win.destroy()
        ttk.Button(win,text="Confirm",command=ok).pack(pady=8)

    def _color_and_measurement(self):
        self.info_lbl.config(text="Select color and enter measurement.")
        win=tk.Toplevel(self); win.title("Color & Measurement"); win.geometry("360x320"); win.grab_set()
        tk.Label(win,text="Color:").pack()
        colors=list(COLORS.keys())
        self.sel_color=tk.StringVar(value=colors[0])
        ttk.OptionMenu(win,self.sel_color,colors[0],*colors).pack()
        tk.Label(win,text="Measurement Type:").pack(pady=(8,0))
        self.measure_type=tk.StringVar(value="Area")
        ttk.Radiobutton(win,text="Area (cm^2)",variable=self.measure_type,value="Area").pack(anchor="w")
        ttk.Radiobutton(win,text="Perimeter (cm)",variable=self.measure_type,value="Perimeter").pack(anchor="w")
        tk.Label(win,text="Value:").pack()
        ent=ttk.Entry(win); ent.insert(0,"100.0"); ent.pack()
        def ok():
            try:
                v=float(ent.get()); assert v>0
            except:
                messagebox.showerror("Error","Enter positive numeric value.",parent=win); return
            self.measure_val=v; win.destroy()
        ttk.Button(win,text="Confirm",command=ok).pack(pady=8)

    def _build_mask(self):
        if self.measure_val is None:
            self.step=2; self._run_step(); return
        if self.input_type=="image":
            hsv=cv2.cvtColor(self.image, cv2.COLOR_BGR2HSV)
            full=None
            for lo,hi in COLORS[self.sel_color.get()]:
                t=cv2.inRange(hsv, lo, hi)
                full = t if full is None else cv2.bitwise_or(full,t)
            k=np.ones((3,3),np.uint8)
            full=cv2.morphologyEx(full, cv2.MORPH_OPEN, k, iterations=2)
            full=cv2.morphologyEx(full, cv2.MORPH_CLOSE, k, iterations=2)
            self.full_color_mask=full
            cnts,_=cv2.findContours(full, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if not cnts:
                messagebox.showerror("Error","No region found for selected color.")
                self.step=2; self._run_step(); return
            big=max(cnts, key=cv2.contourArea)
            main=np.zeros_like(full); cv2.drawContours(main,[big],-1,255,cv2.FILLED)
            self.mask=main
            self.dt_global=cv2.distanceTransform(self.mask, cv2.DIST_L2, 5)
            self.contour=big
            self.main_mask=main.copy()
            self.main_contour=big
            if self.measure_type.get()=="Area":
                A=cv2.contourArea(big)
                if A<=0: messagebox.showerror("Error","Zero area."); self.step=2; self._run_step(); return
                self.px_to_cm=math.sqrt(self.measure_val/A)
            else:
                P=cv2.arcLength(big,True)
                if P<=0: messagebox.showerror("Error","Zero perimeter."); self.step=2; self._run_step(); return
                self.px_to_cm=self.measure_val/P
            base_h=mean_hue(COLORS[self.sel_color.get()])
            self._log(f"px->cm = {self.px_to_cm:.6f} | Main Color Area built.")
        # CSV modunda burada ekstra iş yok; mask/contour zaten _load_csv'de hazır ve px_to_cm de _measurement_csv'de hesaplandı.
        self._redraw()

    def _area_choice(self):
        self.info_lbl.config(text="Choose area creation mode.")
        win=tk.Toplevel(self); win.title("Area Choice"); win.geometry("320x180"); win.grab_set()
        tk.Label(win,text="Select:").pack(pady=6)
        def mine(): self.choice="mine"; win.destroy()
        def whole(): self.choice="whole"; win.destroy()
        ttk.Button(win,text="My Area (click to draw)",command=mine).pack(pady=4)
        if len(self.areas)==0:
            ttk.Button(win,text="Whole (Main Color Area)",command=whole).pack(pady=4)

    def _area_draw(self):
        if not hasattr(self,"choice") or self.choice is None:
            self.step=4; self._run_step(); return
        if self.choice=="whole":
            if self.contour is None:
                messagebox.showerror("Error","Main contour not available."); self.step=3; self._run_step(); return
            area_mask=np.zeros_like(self.mask)
            cv2.drawContours(area_mask,[self.contour],-1,255,cv2.FILLED)
            self.curr_area_mask=area_mask
            self.curr_contour=self.contour.reshape(-1,2)
            self.curr_dt=cv2.distanceTransform(self.curr_area_mask, cv2.DIST_L2, 5)
            self._area_mask_log(self.curr_area_mask, self.curr_contour, "Area (Whole)")
            self._redraw(include_current=True)
            self.step=6; self._run_step(); return
        self.info_lbl.config(text="Click to add vertices. ENTER: close, Backspace: undo, Esc: cancel.")
        self.draw_pts=[]; self.cursor_pt=None
        self._bind_draw(True)
        self._redraw(include_current=True, draw_current_polygon=True)

    def _area_mask_log(self, mask, contour, name_hint):
        per_px=cv2.arcLength(contour.reshape(-1,1,2).astype(np.int32), True)
        ar_px=cv2.contourArea(contour.astype(np.int32))
        per_cm=per_px*self.px_to_cm
        ar_cm2=ar_px*(self.px_to_cm**2)
        M=cv2.moments(contour.astype(np.int32))
        if M["m00"]!=0: cx,cy=M["m10"]/M["m00"], M["m01"]/M["m00"]
        else: cx,cy=centroid(contour)
        self._log(f"[Area Mask] {name_hint} | Perimeter={per_cm:.2f} cm | Area={ar_cm2:.2f} cm^2 | Centroid≈({cx*self.px_to_cm:.2f},{cy*self.px_to_cm:.2f}) cm")

    def _bind_draw(self, enable):
        self._unbind_wizard()
        if enable:
            self.canvas.bind("<Button-1>", self._draw_click)
            self.canvas.bind("<Motion>",   self._draw_move)
            self.canvas.bind("<Key>",      self._draw_key)
            self.canvas.focus_set()
        else:
            self.canvas.unbind("<Button-1>")
            self.canvas.unbind("<Motion>")
            self.canvas.unbind("<Key>")

    def _draw_click(self, ev):
        x,y=self._event_xy_img(ev)
        for a in self.areas:
            if a.mask is not None:
                xi=clamp(int(x),0,a.mask.shape[1]-1); yi=clamp(int(y),0,a.mask.shape[0]-1)
                if a.mask[yi,xi]>0:
                    self._log("Click rejected: overlaps an existing area.")
                    return
        self.draw_pts.append((x,y))
        self._redraw(include_current=True, draw_current_polygon=True)

    def _draw_move(self, ev):
        if not self.draw_pts: return
        self.cursor_pt=self._event_xy_img(ev)
        self._redraw(include_current=True, draw_current_polygon=True)

    def _draw_key(self, ev):
        if ev.keysym=="Return":
            if len(self.draw_pts)<3:
                self._log("Need at least 3 points."); return
            pts=np.array(self.draw_pts, np.int32)
            m=np.zeros(self.image.shape[:2], np.uint8)
            cv2.fillPoly(m,[pts],255)
            self.curr_area_mask=m
            self.curr_contour=pts
            self.curr_dt=cv2.distanceTransform(self.curr_area_mask, cv2.DIST_L2, 5)
            self._bind_draw(False)
            self._area_mask_log(self.curr_area_mask, self.curr_contour, "Area (Draw)")
            self._redraw(include_current=True)
            self.step=6; self._run_step()
        elif ev.keysym in ("BackSpace","Delete"):
            if self.draw_pts:
                self.draw_pts.pop()
                self._redraw(include_current=True, draw_current_polygon=True)
        elif ev.keysym=="Escape":
            self.draw_pts=[]; self.cursor_pt=None
            self._bind_draw(False); self._redraw()

    def _inner_circle(self):
        self.info_lbl.config(text="Largest inscribed circle is shown automatically.")
        if self.curr_area_mask is None:
            if self.curr_contour is not None:
                m=np.zeros(self.image.shape[:2],np.uint8)
                cv2.fillPoly(m,[np.array(self.curr_contour,np.int32)],255)
                self.curr_area_mask=m
                self.curr_dt=cv2.distanceTransform(self.curr_area_mask, cv2.DIST_L2, 5)
            else:
                messagebox.showerror("Error","Area mask missing."); self.step=4; self._run_step(); return
        centers, r, dt = inner_tangent_circle(self.curr_area_mask)
        if not centers or r<=0:
            messagebox.showerror("Error","Cannot find inscribed circle."); self.step=4; self._run_step(); return
        self.curr_dt=dt
        self.centers=centers; self.radius=r
        self.sel_center=centers[0]
        self.canvas.bind("<Button-1>", self._center_click)
        self._redraw(include_current=True, draw_current_center=True)

    def _center_click(self, ev):
        x,y=self._event_xy_img(ev)
        best=None;bd=1e9
        for (mx,my) in self.centers:
            d=math.hypot(mx-x,my-y)
            if d<bd: bd,best=d,(mx,my)
        self.sel_center=best
        self._redraw(include_current=True, draw_current_center=True)

    def _polygon_method(self):
        self.info_lbl.config(text="Choose polygon method.")
        win=tk.Toplevel(self); win.title("Polygon Method"); win.geometry("300x160"); win.grab_set()
        self.poly_method=tk.StringVar(value="Edges")
        ttk.Radiobutton(win,text="By number of edges",variable=self.poly_method,value="Edges").pack(anchor="w", padx=8, pady=4)
        ttk.Radiobutton(win,text="By two points on circle",variable=self.poly_method,value="TwoPoints").pack(anchor="w", padx=8, pady=4)
        ttk.Label(win,text="Inside margin (px):").pack(pady=(6,0))
        sp=ttk.Spinbox(win, from_=0.0, to=5.0, increment=0.5, textvariable=self.inside_margin_px, width=6); sp.pack()
        ttk.Button(win,text="Confirm",command=win.destroy).pack(pady=6)

    def _safe_circumradius(self):
        cx,cy=self.sel_center
        r_dt=self.curr_dt[clamp(int(round(cy)),0,self.curr_dt.shape[0]-1), clamp(int(round(cx)),0,self.curr_dt.shape[1]-1)]
        r0=float(self.radius)
        r1=r0 - float(self.inside_margin_px.get())
        r2=r_dt - 0.5
        return max(1.0, min(r1, r2))

    def _project_on_circle(self, pt, center, r):
        x,y=pt; cx,cy=center
        ang=math.atan2(y-cy, x-cx)
        return (cx + r*math.cos(ang), cy + r*math.sin(ang)), ang

    def _polygon_create(self):
        if self.sel_center is None or not self.radius:
            self.step=6; self._run_step(); return
        if self.poly_method.get()=="Edges":
            self._log("Click a point on the circle.")
            self._poly_clicks=[]; self.canvas.bind("<Button-1>", self._poly_first_click_edges)
        else:
            self._log("Click two points on the circle.")
            self._poly_clicks=[]; self.canvas.bind("<Button-1>", self._poly_two_clicks)
        self._redraw(include_current=True, draw_current_center=True)

    def _poly_first_click_edges(self, ev):
        x,y=self._event_xy_img(ev); cx,cy=self.sel_center
        r=self._safe_circumradius()
        p,_=self._project_on_circle((x,y),(cx,cy), r)
        self._poly_clicks=[p]
        self.canvas.unbind("<Button-1>")
        win=tk.Toplevel(self); win.title("Edges"); win.geometry("280x110"); win.grab_set()
        tk.Label(win,text="Number of edges:").pack()
        ent=ttk.Entry(win); ent.insert(0,"6"); ent.pack()
        def ok():
            try: n=max(4,int(ent.get()))
            except: n=4
            win.destroy(); self._build_polygon_with_n(n, self._poly_clicks[0], r)
        ttk.Button(win,text="Confirm",command=ok).pack(pady=6)

    def _build_polygon_with_n(self, n, first_pt, r):
        cx,cy=self.sel_center
        a0=math.atan2(first_pt[1]-cy, first_pt[0]-cx)
        verts=[]
        for i in range(n):
            th=a0 + i*(2*math.pi/n)
            verts.append((cx + r*math.cos(th), cy + r*math.sin(th)))
        verts=[(int(round(x)), int(round(y))) for (x,y) in verts]
        self._finalize_polygon(verts, method="by_edges", n=n, angle=a0, circ_center=(cx,cy), circ_r=r, circ_flag=True)

    def _poly_two_clicks(self, ev):
        x,y=self._event_xy_img(ev); cx,cy=self.sel_center
        r=self._safe_circumradius()
        p,ang=self._project_on_circle((x,y),(cx,cy), r)
        self._poly_clicks.append((p,ang,r))
        if len(self._poly_clicks)==2:
            self.canvas.unbind("<Button-1>")
            (p1,a1,r1),(p2,a2,r2)=self._poly_clicks
            r=min(r1,r2)
            d=a2-a1
            if d<0: d+=2*math.pi
            if d <= (2*math.pi - d): sgn=1; delta=d
            else: sgn=-1; delta=2*math.pi - d
            n_approx=max(4, round((4*math.pi)/delta))
            verts=[]
            for i in range(n_approx):
                th=a1+sgn*i*(2*math.pi/n_approx)
                verts.append((cx + r*math.cos(th), cy + r*math.sin(th)))
            verts=[(int(round(x)), int(round(y))) for (x,y) in verts]
            self._finalize_polygon(verts, method="by_twopoints", n=n_approx, angle=a1, circ_center=(cx,cy), circ_r=r, circ_flag=True)

    def _finalize_polygon(self, vertices_px, method, n, angle, circ_center, circ_r, circ_flag):
        H,W=self.image.shape[:2]
        if self.curr_area_mask is None:
            if self.curr_contour is not None:
                m=np.zeros((H,W),np.uint8); cv2.fillPoly(m,[np.array(self.curr_contour,np.int32)],255)
                self.curr_area_mask=m
                self.curr_dt=cv2.distanceTransform(self.curr_area_mask, cv2.DIST_L2, 5)
            else:
                messagebox.showerror("Error","Area mask missing."); self.step=4; self._run_step(); return
        cx,cy=circ_center
        dloc=self.curr_dt[clamp(int(round(cy)),0,self.curr_dt.shape[0]-1), clamp(int(round(cx)),0,self.curr_dt.shape[1]-1)]
        if circ_r > dloc + 1e-6:
            messagebox.showerror("Error","Geometric test failed: polygon radius exceeds local clearance."); return
        cand_m=polygon_mask(vertices_px, (H,W))
        for a in self.areas:
            if a.polygon_mask is None: continue
            if eroded_overlap(cand_m, a.polygon_mask, iters=1):
                messagebox.showerror("Error", f"Overlaps with '{a.name}' polygon."); return
        self.area_counter+=1
        area_name=f"Area {self.area_counter}"
        base_h = mean_hue(COLORS[self.sel_color.get()]) if (self.input_type=="image") else 105
        cont = self.curr_contour if self.curr_contour is not None else np.array(vertices_px, np.int32)
        new = Area(self.area_counter, area_name, self.curr_area_mask.copy(), np.array(cont,np.int32), base_h)
        new.center=self.sel_center; new.rmax=self.radius
        new.polygon={"method":method, "n":int(n), "angle":float(angle),
                     "vertices_px":[(int(x),int(y)) for (x,y) in vertices_px],
                     "vertices_cm":None, "perimeter_cm":None, "area_cm2":None,
                     "is_circ":bool(circ_flag), "circ_center_px":(float(cx),float(cy)), "circ_radius_px":float(circ_r)}
        new.polygon_mask=cand_m
        self._update_metrics_and_inclusion(new, log=True)
        self.areas.append(new)
        self.choice=None
        self.curr_area_mask=None; self.curr_contour=None; self.curr_dt=None
        self.draw_pts=[]; self.cursor_pt=None
        self.centers=[]; self.sel_center=None; self.radius=None
        self._redraw()
        if messagebox.askyesno("Area","Create another area?"):
            self.step=4; self._run_step()
        else:
            self.step=9; self._run_step()

    def _update_metrics_and_inclusion(self, area_obj, log=False):
        verts=area_obj.polygon["vertices_px"]; scale=self.px_to_cm
        per_px=poly_perimeter(verts); ar_px=poly_area(verts)
        area_obj.polygon["perimeter_cm"]=per_px*scale
        area_obj.polygon["area_cm2"]=ar_px*(scale**2)
        area_obj.polygon["vertices_cm"]=[(x*scale,y*scale) for (x,y) in verts]
        if self.prox_ref.get()=="Main Color Area" and self.main_mask is not None:
            ref_mask=self.main_mask; ref_cont=self.main_contour
        else:
            ref_mask=area_obj.mask; ref_cont=area_obj.contour
        inter=cv2.bitwise_and(polygon_mask(verts,self.image.shape[:2]), ref_mask)
        touches=cv2.countNonZero(inter)>0
        cntP,_=cv2.findContours(polygon_mask(verts,self.image.shape[:2]), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cP=max(cntP, key=cv2.contourArea) if cntP else None
        dist_px=0.0 if touches else (min_contour_distance(ref_cont.reshape(-1,2), cP.reshape(-1,2)) if cP is not None else float('inf'))
        area_obj.proximity_px=dist_px; area_obj.proximity_cm=(dist_px*scale if math.isfinite(dist_px) else None)
        if not self.exclude_if_far.get():
            area_obj.included=True; area_obj.exclude_reason=None
        else:
            thr=self.proximity_px.get()
            if touches or (math.isfinite(dist_px) and dist_px<=thr):
                area_obj.included=True; area_obj.exclude_reason=None
            else:
                area_obj.included=False; area_obj.exclude_reason=f"far_from_ref ({dist_px:.1f}px > {thr}px)"
        if log:
            cen=centroid(verts); cen_cm=(cen[0]*scale, cen[1]*scale)
            inc="✓" if area_obj.included else "✗"
            self._log(f"{area_obj.name} | n={len(verts)} | Perimeter={area_obj.polygon['perimeter_cm']:.2f} cm | Area={area_obj.polygon['area_cm2']:.2f} cm^2 | Centroid≈({cen_cm[0]:.2f},{cen_cm[1]:.2f}) cm | included={inc} | prox={area_obj.proximity_px:.1f}px (~{(area_obj.proximity_cm or 0):.2f}cm) {('['+area_obj.exclude_reason+']') if area_obj.exclude_reason else ''}")

    def _recalc_inclusions(self):
        for a in self.areas:
            if a.polygon: self._update_metrics_and_inclusion(a, log=False)
        self._redraw()

    def _summary_and_export(self):
        self.info_lbl.config(text="Summary and export.")
        if not self.areas: self._log("No data."); return
        self._log("========== SUMMARY ==========")
        src=os.path.basename(self.path) if self.path else "-"
        self._log(f"Source: {src} | px->cm={self.px_to_cm:.6f} | Proximity={self.proximity_px.get()}px | Ref={self.prox_ref.get()} | Filter={'ON' if self.exclude_if_far.get() else 'OFF'}")
        self._log(f"Constraints: local={'ON' if self.enforce_local.get() else 'OFF'} | global={'ON' if self.enforce_global.get() else 'OFF'}")
        totA=0.0; totP=0.0
        for a in self.areas:
            inc="✓" if a.included else "✗"
            prox=f"{a.proximity_px:.1f}px (~{(a.proximity_cm or 0):.2f}cm)"
            self._log(f"- {a.name} | included={inc} | n={a.polygon['n']} | Perimeter={a.polygon['perimeter_cm']:.2f} cm | Area={a.polygon['area_cm2']:.2f} cm^2 | prox={prox} {('['+a.exclude_reason+']') if a.exclude_reason else ''}")
            vc=a.polygon["vertices_cm"]; L=[]
            for i in range(len(vc)):
                x1,y1=vc[i]; x2,y2=vc[(i+1)%len(vc)]
                L.append(math.hypot(x2-x1,y2-y1))
            self._log("  Edges: " + ", ".join([f"e{i+1}={l:.2f}cm" for i,l in enumerate(L)]))
            if a.included: totA+=a.polygon["area_cm2"]; totP+=a.polygon["perimeter_cm"]
        self._log(f"TOTAL (included=✓): Area={totA:.2f} cm^2 | Perimeter={totP:.2f} cm")
        self._log("=============================")
        win=tk.Toplevel(self); win.title("Export"); win.geometry("420x260"); win.grab_set()
        ttk.Button(win,text="Save annotated PNG",command=self._save_png).pack(pady=6)
        ttk.Button(win,text="Export XLSX (pretty) or CSV fallback",command=self._export_tables).pack(pady=6)
        ttk.Button(win,text="Export JSON summary",command=self._export_json).pack(pady=6)

    def _save_png(self):
        img=self._compose_overlay()
        if img is None: return
        for a in self.areas:
            if not a.polygon: continue
            verts=a.polygon["vertices_px"]
            col=(160,160,160) if not a.included else (255,255,255)
            for i,(x,y) in enumerate(verts, start=1):
                cv2.putText(img, str(i), (int(x)+3,int(y)-3), cv2.FONT_HERSHEY_SIMPLEX, 0.45, col, 1, cv2.LINE_AA)
            vc=a.polygon["vertices_cm"]
            for i,(x1,y1) in enumerate(verts):
                x2,y2=verts[(i+1)%len(verts)]
                L=math.hypot(vc[(i+1)%len(vc)][0]-vc[i][0], vc[(i+1)%len(vc)][1]-vc[i][1])
                mx,my=int((x1+x2)/2), int((y1+y2)/2)
                cv2.putText(img, f"{L:.2f}cm", (mx,my), cv2.FONT_HERSHEY_SIMPLEX, 0.45, col, 1, cv2.LINE_AA)
        path=filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG","*.png")], title="Save")
        if not path: return
        if os.path.exists(path): os.remove(path)
        cv2.imwrite(path, img); messagebox.showinfo("OK", os.path.basename(path)+" saved.")

    def _export_tables(self):
        folder=filedialog.askdirectory(title="Select folder")
        if not folder: return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            wb=Workbook()
            def autosize(ws):
                widths={}
                for row in ws.rows:
                    for cell in row:
                        v=str(cell.value) if cell.value is not None else ""
                        widths[cell.column_letter]=max(widths.get(cell.column_letter,0), len(v)+2)
                for col,w in widths.items():
                    ws.column_dimensions[col].width=min(60,w)
            ws1=wb.active; ws1.title="Areas"
            ws1.append(["area_id","area_name","perimeter_cm","area_cm2","centroid_px_x","centroid_px_y","centroid_cm_x","centroid_cm_y"])
            for a in self.areas:
                cont=a.contour.astype(np.int32)
                per_px=cv2.arcLength(cont.reshape(-1,1,2), True)
                ar_px=cv2.contourArea(cont)
                per_cm=per_px*self.px_to_cm; ar_cm2=ar_px*(self.px_to_cm**2)
                M=cv2.moments(cont)
                if M["m00"]!=0: cx,cy=M["m10"]/M["m00"], M["m01"]/M["m00"]
                else:
                    c=centroid(cont.reshape(-1,2)); cx,cy=c[0],c[1]
                ws1.append([a.id,a.name,round(per_cm,3),round(ar_cm2,3),int(round(cx)),int(round(cy)),round(cx*self.px_to_cm,3),round(cy*self.px_to_cm,3)])
            for c in ws1[1]: c.font=Font(bold=True)
            autosize(ws1)
            ws2=wb.create_sheet("Polygons")
            ws2.append(["area_id","area_name","included","reason","n_vertices","perimeter_cm","area_cm2","centroid_px_x","centroid_px_y","centroid_cm_x","centroid_cm_y","proximity_px","proximity_cm","is_circ","circ_center_x","circ_center_y","circ_radius_px"])
            for a in self.areas:
                cen=centroid(a.polygon["vertices_px"])
                circ=a.polygon.get("is_circ",False)
                cc=a.polygon.get("circ_center_px",(None,None))
                rr=a.polygon.get("circ_radius_px",None)
                ws2.append([a.id,a.name,1 if a.included else 0,a.exclude_reason if a.exclude_reason else "",len(a.polygon["vertices_px"]),round(a.polygon["perimeter_cm"],3),round(a.polygon["area_cm2"],3),int(round(cen[0])),int(round(cen[1])),round(cen[0]*self.px_to_cm,3),round(cen[1]*self.px_to_cm,3),round(a.proximity_px if a.proximity_px is not None else 0.0,3),round(a.proximity_cm if a.proximity_cm is not None else 0.0,3),1 if circ else 0, None if cc[0] is None else round(cc[0],3), None if cc[1] is None else round(cc[1],3), None if rr is None else round(rr,3)])
            for c in ws2[1]: c.font=Font(bold=True)
            autosize(ws2)
            ws3=wb.create_sheet("Edges")
            ws3.append(["area_id","area_name","edge_idx","length_cm","p1_px_x","p1_px_y","p2_px_x","p2_px_y"])
            for a in self.areas:
                vp=a.polygon["vertices_px"]; vc=a.polygon["vertices_cm"]
                for i,(x1,y1) in enumerate(vp):
                    x2,y2=vp[(i+1)%len(vp)]
                    L=math.hypot(vc[(i+1)%len(vc)][0]-vc[i][0], vc[(i+1)%len(vc)][1]-vc[i][1])
                    ws3.append([a.id,a.name,i+1,round(L,3),int(x1),int(y1),int(x2),int(y2)])
            for c in ws3[1]: c.font=Font(bold=True)
            autosize(ws3)
            ws4=wb.create_sheet("Vertices")
            ws4.append(["area_id","area_name","vertex_idx","x_px","y_px","x_cm","y_cm"])
            for a in self.areas:
                for idx,((xpx,ypx),(xcm,ycm)) in enumerate(zip(a.polygon["vertices_px"], a.polygon["vertices_cm"]), start=1):
                    ws4.append([a.id,a.name,idx,int(xpx),int(ypx),round(xcm,3),round(ycm,3)])
            for c in ws4[1]: c.font=Font(bold=True)
            autosize(ws4)
            ts=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            xlsx_path=os.path.join(folder,f"polygons_{ts}.xlsx")
            wb.save(xlsx_path)
            messagebox.showinfo("OK", os.path.basename(xlsx_path)+" saved.")
        except Exception:
            ts=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            areas_p=os.path.join(folder,f"areas_{ts}.csv")
            polys_p=os.path.join(folder,f"polygons_{ts}.csv")
            edges_p=os.path.join(folder,f"edges_{ts}.csv")
            verts_p=os.path.join(folder,f"vertices_{ts}.csv")
            with open(areas_p,"w",newline="",encoding="utf-8") as f:
                w=csv.writer(f); w.writerow(["area_id","area_name","perimeter_cm","area_cm2","centroid_px_x","centroid_px_y","centroid_cm_x","centroid_cm_y"])
                for a in self.areas:
                    cont=a.contour.astype(np.int32)
                    per_px=cv2.arcLength(cont.reshape(-1,1,2), True)
                    ar_px=cv2.contourArea(cont)
                    per_cm=per_px*self.px_to_cm; ar_cm2=ar_px*(self.px_to_cm**2)
                    M=cv2.moments(cont)
                    if M["m00"]!=0: cx,cy=M["m10"]/M["m00"], M["m01"]/M["m00"]
                    else:
                        c=centroid(cont.reshape(-1,2)); cx,cy=c[0],c[1]
                    w.writerow([a.id,a.name,f"{per_cm:.3f}",f"{ar_cm2:.3f}",int(round(cx)),int(round(cy)),f"{cx*self.px_to_cm:.3f}",f"{cy*self.px_to_cm:.3f}"])
            with open(polys_p,"w",newline="",encoding="utf-8") as f:
                w=csv.writer(f); w.writerow(["area_id","area_name","included","reason","n_vertices","perimeter_cm","area_cm2","centroid_px_x","centroid_px_y","centroid_cm_x","centroid_cm_y","proximity_px","proximity_cm","is_circ","circ_center_x","circ_center_y","circ_radius_px"])
                for a in self.areas:
                    cen=centroid(a.polygon["vertices_px"])
                    circ=a.polygon.get("is_circ",False)
                    cc=a.polygon.get("circ_center_px",(None,None))
                    rr=a.polygon.get("circ_radius_px",None)
                    w.writerow([a.id,a.name,1 if a.included else 0,a.exclude_reason if a.exclude_reason else "",len(a.polygon["vertices_px"]),f"{a.polygon['perimeter_cm']:.3f}",f"{a.polygon['area_cm2']:.3f}",int(round(cen[0])),int(round(cen[1])),f"{cen[0]*self.px_to_cm:.3f}",f"{cen[1]*self.px_to_cm:.3f}",f"{(a.proximity_px if a.proximity_px is not None else 0.0):.3f}",f"{(a.proximity_cm if a.proximity_cm is not None else 0.0):.3f}",1 if circ else 0, "" if cc[0] is None else f"{cc[0]:.3f}", "" if cc[1] is None else f"{cc[1]:.3f}", "" if rr is None else f"{rr:.3f}"])
            with open(edges_p,"w",newline="",encoding="utf-8") as f:
                w=csv.writer(f); w.writerow(["area_id","area_name","edge_idx","length_cm","p1_px_x","p1_px_y","p2_px_x","p2_px_y"])
                for a in self.areas:
                    vp=a.polygon["vertices_px"]; vc=a.polygon["vertices_cm"]
                    for i,(x1,y1) in enumerate(vp):
                        x2,y2=vp[(i+1)%len(vp)]
                        L=math.hypot(vc[(i+1)%len(vc)][0]-vc[i][0], vc[(i+1)%len(vc)][1]-vc[i][1])
                        w.writerow([a.id,a.name,i+1,f"{L:.3f}",int(x1),int(y1),int(x2),int(y2)])
            with open(verts_p,"w",newline="",encoding="utf-8") as f:
                w=csv.writer(f); w.writerow(["area_id","area_name","vertex_idx","x_px","y_px","x_cm","y_cm"])
                for a in self.areas:
                    for idx,((xpx,ypx),(xcm,ycm)) in enumerate(zip(a.polygon["vertices_px"], a.polygon["vertices_cm"]), start=1):
                        w.writerow([a.id,a.name,idx,int(xpx),int(ypx),f"{xcm:.3f}",f"{ycm:.3f}"])
            messagebox.showinfo("CSV saved","openpyxl not available. CSV files were saved instead.")

    def _export_json(self):
        if not self.areas: return
        path=filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON","*.json")], title="Save")
        if not path: return
        data={"source": os.path.basename(self.path) if self.path else None,
              "px_to_cm": self.px_to_cm,
              "proximity_px_threshold": self.proximity_px.get(),
              "proximity_reference": self.prox_ref.get(),
              "proximity_filter": bool(self.exclude_if_far.get()),
              "constraints":{"local": bool(self.enforce_local.get()), "global": bool(self.enforce_global.get())},
              "areas":[]}
        for a in self.areas:
            cont=a.contour.astype(np.int32)
            per_px=cv2.arcLength(cont.reshape(-1,1,2), True)
            ar_px=cv2.contourArea(cont)
            per_cm=per_px*self.px_to_cm; ar_cm2=ar_px*(self.px_to_cm**2)
            M=cv2.moments(cont)
            if M["m00"]!=0: cx,cy=M["m10"]/M["m00"], M["m01"]/M["m00"]
            else:
                c=centroid(cont.reshape(-1,2)); cx,cy=c[0],c[1]
            ent={"id":a.id,"name":a.name,"included":bool(a.included),"exclude_reason":a.exclude_reason,
                 "mask_metrics":{"perimeter_cm": float(per_cm), "area_cm2": float(ar_cm2),
                                 "centroid_px":[float(cx),float(cy)],
                                 "centroid_cm":[float(cx*self.px_to_cm), float(cy*self.px_to_cm)]},
                 "polygon":{}}
            if a.polygon:
                ent["polygon"]={
                    "n": int(a.polygon["n"]),
                    "perimeter_cm": float(a.polygon["perimeter_cm"]),
                    "area_cm2": float(a.polygon["area_cm2"]),
                    "centroid_px": [float(v) for v in centroid(a.polygon["vertices_px"])],
                    "centroid_cm": [float(v*self.px_to_cm) for v in centroid(a.polygon["vertices_px"])],
                    "proximity_px": float(a.proximity_px) if a.proximity_px is not None else None,
                    "proximity_cm": float(a.proximity_cm) if a.proximity_cm is not None else None,
                    "is_circ": bool(a.polygon.get("is_circ",False)),
                    "circ_center_px": list(a.polygon.get("circ_center_px",(None,None))) if a.polygon.get("is_circ",False) else None,
                    "circ_radius_px": float(a.polygon.get("circ_radius_px",0.0)) if a.polygon.get("is_circ",False) else None,
                    "vertices":[{"idx":i+1,"px":[int(p[0]),int(p[1])],
                                 "cm":[float(a.polygon["vertices_cm"][i][0]), float(a.polygon["vertices_cm"][i][1])]}
                                for i,p in enumerate(a.polygon["vertices_px"])],
                    "edges":[{"idx":i+1,
                              "p1_px":[int(a.polygon["vertices_px"][i][0]), int(a.polygon["vertices_px"][i][1])],
                              "p2_px":[int(a.polygon["vertices_px"][(i+1)%len(a.polygon['vertices_px'])][0]),
                                       int(a.polygon["vertices_px"][(i+1)%len(a.polygon['vertices_px'])][1])],
                              "length_cm": float(math.hypot(
                                  a.polygon["vertices_cm"][(i+1)%len(a.polygon['vertices_cm'])][0] - a.polygon["vertices_cm"][i][0],
                                  a.polygon["vertices_cm"][(i+1)%len(a.polygon['vertices_cm'])][1] - a.polygon["vertices_cm"][i][1]
                              ))} for i in range(len(a.polygon["vertices_px"]))]}
            data["areas"].append(ent)
        with open(path,"w",encoding="utf-8") as f:
            json.dump(data,f,ensure_ascii=False,indent=2)
        messagebox.showinfo("OK", os.path.basename(path)+" saved.")

    def _wizard_toggle(self):
        if not self.wizard_on.get(): self._unbind_wizard()
        self._update_wizard_bindings(); self._redraw()

    def _unbind_wizard(self):
        self.canvas.unbind("<Button-1>")
        self.canvas.unbind("<B1-Motion>")
        self.canvas.unbind("<ButtonRelease-1>")
        self.canvas.unbind("<MouseWheel>")
        self.canvas.unbind("<Key>")
        self.canvas.unbind("<KeyPress-f>")
        self.canvas.unbind("<KeyRelease-f>")
        self.canvas.unbind("<Button-4>")
        self.canvas.unbind("<Button-5>")

    def _update_wizard_bindings(self):
        if self.step in (5,6,7,8):
            self.btn_prepare.state(["disabled"]); self.btn_commit.state(["disabled"]); self.btn_cancel.state(["disabled"])
            return
        self._unbind_wizard()
        if self.merge_state=="prep":
            self.btn_prepare.state(["disabled"]); self.btn_commit.state(["!disabled"]); self.btn_cancel.state(["!disabled"])
            self.canvas.bind("<Button-1>", self._merge_click)
            self.canvas.bind("<B1-Motion>", self._merge_drag)
            self.canvas.bind("<ButtonRelease-1>", self._merge_release)
            self.canvas.bind("<MouseWheel>", self._merge_wheel)
            self.canvas.bind("<Key>", self._merge_key)
            self.canvas.bind("<KeyPress-f>", lambda e: self._set_snap(True))
            self.canvas.bind("<KeyRelease-f>", lambda e: self._set_snap(False))
            self.canvas.bind("<Button-4>", lambda e: self._merge_rotate(+1))
            self.canvas.bind("<Button-5>", lambda e: self._merge_rotate(-1))
            return
        if not self.wizard_on.get():
            self.btn_prepare.state(["disabled"]); self.btn_commit.state(["disabled"]); self.btn_cancel.state(["disabled"])
            return
        self.btn_prepare.state(["!disabled"]); self.btn_commit.state(["disabled"]); self.btn_cancel.state(["disabled"])
        self.canvas.bind("<Button-1>", self._wiz_click)
        self.canvas.bind("<B1-Motion>", self._wiz_drag)
        self.canvas.bind("<ButtonRelease-1>", self._wiz_release)
        self.canvas.bind("<MouseWheel>", self._wiz_wheel)
        self.canvas.bind("<Key>", self._wiz_key)
        self.canvas.bind("<KeyPress-f>", lambda e: self._set_snap(True))
        self.canvas.bind("<KeyRelease-f>", lambda e: self._set_snap(False))
        self.canvas.bind("<Button-4>", lambda e: self._wiz_rotate(+1))
        self.canvas.bind("<Button-5>", lambda e: self._wiz_rotate(-1))
        self._log("Wizard ON: left-click select/drag; wheel rotate; hold F to snap; [ / ] step.")

    def _set_snap(self, val):
        self.snap_active=val

    def _wiz_click(self, ev):
        if not self.wizard_on.get(): return
        x,y=self._event_xy_img(ev)
        self.selected_area=None
        for a in reversed(self.areas):
            if a.polygon and point_in_poly((x,y), a.polygon["vertices_px"]):
                self.selected_area=a; break
        if not self.selected_area:
            self._log("No polygon selected (click inside one)."); return
        self.dragging=True; self.drag_start=(x,y)
        self.wiz_orig=[tuple(v) for v in self.selected_area.polygon["vertices_px"]]
        self.wiz_pivot=centroid(self.wiz_orig)
        self.wiz_dx=0; self.wiz_dy=0; self.wiz_angle=0.0
        self.drag_preview=None; self.drag_valid=False
        self.canvas.focus_set(); self._redraw()

    def _geom_inside_local(self, cand):
        if not self.selected_area: return True
        if not self.enforce_local.get(): return True
        if not self.selected_area.polygon.get("is_circ",False): return None
        cx,cy=centroid(cand)
        d=self.selected_area.dt[clamp(int(round(cy)),0,self.selected_area.dt.shape[0]-1), clamp(int(round(cx)),0,self.selected_area.dt.shape[1]-1)]
        r=float(self.selected_area.polygon.get("circ_radius_px",0.0))
        return r <= d - 0.25

    def _geom_inside_global(self, cand):
        if not self.enforce_global.get(): return True
        if self.dt_global is None: return None
        if not self.selected_area or not self.selected_area.polygon.get("is_circ",False): return None
        cx,cy=centroid(cand)
        d=self.dt_global[clamp(int(round(cy)),0,self.dt_global.shape[0]-1), clamp(int(round(cx)),0,self.dt_global.shape[1]-1)]
        r=float(self.selected_area.polygon.get("circ_radius_px",0.0))
        return r <= d - 0.25

    def _check_constraints(self, cand, ignore=None):
        H,W=self.image.shape[:2]
        ok_local=True; ok_global=True
        gi=self._geom_inside_local(cand)
        if gi is None:
            if self.enforce_local.get() and self.selected_area and self.selected_area.mask is not None:
                ok_local = cv2.countNonZero(cv2.bitwise_and(polygon_mask(cand,(H,W)), cv2.bitwise_not(self.selected_area.mask)))==0
        else:
            ok_local = gi
        gg=self._geom_inside_global(cand)
        if gg is None:
            if self.enforce_global.get() and self.mask is not None:
                ok_global = cv2.countNonZero(cv2.bitwise_and(polygon_mask(cand,(H,W)), cv2.bitwise_not(self.mask)))==0
        else:
            ok_global = gg
        cand_m=polygon_mask(cand,(H,W))
        ok_inter=True
        for a in self.areas:
            if a is self.selected_area or a is ignore: continue
            if a.polygon_mask is None: continue
            if eroded_overlap(cand_m, a.polygon_mask, iters=1):
                ok_inter=False; break
        return ok_local and ok_global and ok_inter

    def _wiz_current_preview(self):
        rotated=rotate_points(self.wiz_orig, self.wiz_pivot, self.wiz_angle)
        cand=[(x+self.wiz_dx, y+self.wiz_dy) for (x,y) in rotated]
        if self.snap_active:
            targets=[]
            for a in self.areas:
                if a is self.selected_area or not a.polygon: continue
                targets += a.polygon["vertices_px"]
            best=None; bd=1e9
            for (cx,cy) in cand:
                for (tx,ty) in targets:
                    d=math.hypot(tx-cx,ty-cy)
                    if d<bd and d<=self.snap_px:
                        bd=d; best=(tx-cx, ty-cy)
            if best is not None:
                cand=[(px+best[0],py+best[1]) for (px,py) in cand]
        return cand

    def _wiz_drag(self, ev):
        if not (self.wizard_on.get() and self.dragging and self.selected_area): return
        x,y=self._event_xy_img(ev)
        self.wiz_dx=x-self.drag_start[0]; self.wiz_dy=y-self.drag_start[1]
        cand=self._wiz_current_preview()
        self.drag_preview=cand; self.drag_valid=self._check_constraints(cand)
        self._redraw()

    def _wiz_release(self, ev):
        if not (self.wizard_on.get() and self.dragging and self.selected_area): return
        cand=self._wiz_current_preview()
        if cand is not None and self._check_constraints(cand):
            a=self.selected_area
            a.polygon["vertices_px"]=[(int(round(x)),int(round(y))) for (x,y) in cand]
            a.polygon_mask=polygon_mask(a.polygon["vertices_px"], self.image.shape[:2])
            if a.polygon.get("is_circ",False):
                cen=centroid(a.polygon["vertices_px"])
                a.polygon["circ_center_px"]=(float(cen[0]), float(cen[1]))
            self._update_metrics_and_inclusion(a, log=True)
        else:
            self._log("Move/rotate cancelled due to constraints.")
        self.dragging=False; self.drag_start=None
        self.wiz_orig=None; self.wiz_pivot=None; self.wiz_dx=0; self.wiz_dy=0; self.wiz_angle=0.0
        self.drag_preview=None; self.drag_valid=False
        self._redraw()

    def _wiz_wheel(self, ev):
        if not (self.wizard_on.get() and self.dragging): return
        direction=+1 if ev.delta>0 else -1
        self._wiz_rotate(direction)

    def _wiz_rotate(self, direction):
        if not (self.wizard_on.get() and self.dragging): return
        step=math.radians(self.rotate_step.get())*(1 if direction>0 else -1)
        self.wiz_angle+=step
        cand=self._wiz_current_preview()
        self.drag_preview=cand; self.drag_valid=self._check_constraints(cand)
        self._redraw()

    def _wiz_key(self, ev):
        k=ev.keysym.lower()
        if k in ("q","e"):
            self._wiz_rotate(+1 if k=="e" else -1)
        elif k=="bracketleft":
            self.rotate_step.set(clamp(self.rotate_step.get()-0.5,0.1,15.0)); self._log(f"Step: {self.rotate_step.get():.1f}°")
        elif k=="bracketright":
            self.rotate_step.set(clamp(self.rotate_step.get()+0.5,0.1,15.0)); self._log(f"Step: {self.rotate_step.get():.1f}°")

    def _edge_list(self, verts):
        return [((verts[i][0],verts[i][1]), (verts[(i+1)%len(verts)][0],verts[(i+1)%len(verts)][1])) for i in range(len(verts))]

    def _best_edge_pair(self, Av, Bv):
        Aedges=self._edge_list(Av); Bedges=self._edge_list(Bv)
        best=None; best_score=1e18; best_data=None
        for ia,(a1,a2) in enumerate(Aedges):
            ax1,ay1=a1; ax2,ay2=a2
            va=(ax2-ax1, ay2-ay1)
            lenA=math.hypot(va[0],va[1])
            if lenA<1e-6: continue
            angA=math.atan2(va[1],va[0])
            for ib,(b1,b2) in enumerate(Bedges):
                bx1,by1=b1; bx2,by2=b2
                vb=(bx2-bx1, by2-by1)
                lenB=math.hypot(vb[0],vb[1])
                if lenB<1e-6: continue
                angB=math.atan2(vb[1],vb[0])
                rot = angA - (angB + math.pi)
                Bc=centroid(Bv)
                Brot=rotate_points(Bv, Bc, rot)
                b1r=rotate_points([b1], Bc, rot)[0]
                b2r=rotate_points([b2], Bc, rot)[0]
                tx=ax1 - b2r[0]; ty=ay1 - b2r[1]
                b1t=(b1r[0]+tx, b1r[1]+ty)
                b2t=(b2r[0]+tx, b2r[1]+ty)
                len_diff=abs(lenA - math.hypot(b2t[0]-b1t[0], b2t[1]-b1t[1]))
                other_end_err=math.hypot(ax2 - b1t[0], ay2 - b1t[1])
                score = other_end_err + 3.0*len_diff
                if score<best_score:
                    best_score=score; best=(ia,ib); best_data=(Brot, tx, ty, rot, (ax1,ay1,ax2,ay2))
        return best, best_data

    def _merge_prepare_prompt(self):
        if not self.wizard_on.get():
            messagebox.showinfo("Merge","Turn on Wizard and select a base polygon first."); return
        if self.merge_state=="prep":
            messagebox.showinfo("Merge","Alignment already active."); return
        if not self.selected_area:
            messagebox.showinfo("Merge","Click inside a polygon to select it as base, then try again."); return
        candidates=[a for a in self.areas if a is not self.selected_area and a.polygon is not None]
        if not candidates:
            messagebox.showinfo("Merge","No other polygons to merge."); return
        win=tk.Toplevel(self); win.title("Prepare Merge"); win.geometry("360x180"); win.grab_set()
        tk.Label(win,text=f"Base: {self.selected_area.name}\nMerge with:").pack(pady=6)
        names=[f"{a.id}: {a.name}" for a in candidates]
        sel=tk.StringVar(value=names[0])
        ttk.OptionMenu(win, sel, names[0], *names).pack()
        def ok():
            idx=names.index(sel.get())
            win.destroy()
            self._prepare_alignment(self.selected_area, candidates[idx])
        ttk.Button(win,text="Prepare",command=ok).pack(pady=8)

    def _prepare_alignment(self, base, move):
        Av=list(base.polygon["vertices_px"])
        Bv=list(move.polygon["vertices_px"])
        if signed_area(Av)<0: Av=list(reversed(Av))
        if signed_area(Bv)<0: Bv=list(reversed(Bv))
        best,bdata=self._best_edge_pair(Av,Bv)
        if best is None:
            messagebox.showerror("Merge","No compatible edges found."); return
        ia, ib = best
        Brot, tx, ty, rot, aedge = bdata
        Btf=[(x+tx,y+ty) for (x,y) in Brot]
        self.merge_state="prep"
        self.merge_base=base
        self.merge_move=move
        self.merge_best=(ia, ib)
        self.merge_init_vertices=Btf
        ax1,ay1,ax2,ay2=aedge
        self.merge_pivot=((ax1+ax2)/2.0,(ay1+ay2)/2.0)
        self.merge_dx=0.0; self.merge_dy=0.0; self.merge_angle=0.0
        self.canvas.focus_set()
        self._update_wizard_bindings()
        self._log(f"Alignment prepared: {base.name} ↔ {move.name} | edges A#{ia+1}–B#{ib+1}")
        self._redraw()

    def _merge_current_candidate(self):
        if self.merge_state!="prep" or self.merge_init_vertices is None: return None
        cand=rotate_points(self.merge_init_vertices, self.merge_pivot, self.merge_angle)
        cand=[(x+self.merge_dx, y+self.merge_dy) for (x,y) in cand]
        if self.snap_active:
            targets=self.merge_base.polygon["vertices_px"]
            best=None; bd=1e9
            for (cx,cy) in cand:
                for (tx,ty) in targets:
                    d=math.hypot(tx-cx,ty-cy)
                    if d<bd and d<=self.snap_px:
                        bd=d; best=(tx-cx, ty-cy)
            if best is not None:
                cand=[(px+best[0],py+best[1]) for (px,py) in cand]
        return cand

    def _merge_metrics(self, cand):
        ia,ib=self.merge_best
        Av=list(self.merge_base.polygon["vertices_px"])
        Bv=list(cand)
        if signed_area(Av)<0: Av=list(reversed(Av))
        if signed_area(Bv)<0: Bv=list(reversed(Bv))
        ax1,ay1=Av[ia]; ax2,ay2=Av[(ia+1)%len(Av)]
        bx1,by1=Bv[ib]; bx2,by2=Bv[(ib+1)%len(Bv)]
        vA=(ax2-ax1, ay2-ay1)
        vB=(bx2-bx1, by2-by1)
        angA=math.atan2(vA[1],vA[0])
        angB=math.atan2(vB[1],vB[0])+math.pi
        da=(angA-angB+math.pi)%(2*math.pi)-math.pi
        n=(-vA[1], vA[0]); ln=math.hypot(n[0],n[1]); n=(n[0]/ln, n[1]/ln)
        mxA=((ax1+ax2)/2.0, (ay1+ay2)/2.0)
        mxB=((bx1+bx2)/2.0, (by1+by2)/2.0)
        off=abs((mxB[0]-mxA[0])*n[0] + (mxB[1]-mxA[1])*n[1])
        eg1=math.hypot(ax1-bx2, ay1-by2)
        eg2=math.hypot(ax2-bx1, ay2-by1)
        return abs(math.degrees(da)), off, max(eg1,eg2)

    def _merge_click(self, ev):
        if self.merge_state!="prep": return
        self.merge_start=self._event_xy_img(ev)

    def _merge_drag(self, ev):
        if self.merge_state!="prep": return
        x,y=self._event_xy_img(ev)
        sx,sy=self.merge_start
        self.merge_dx += (x-sx); self.merge_dy += (y-sy)
        self.merge_start=(x,y)
        self._redraw()

    def _merge_release(self, ev):
        if self.merge_state!="prep": return
        self._redraw()

    def _merge_wheel(self, ev):
        if self.merge_state!="prep": return
        direction=+1 if ev.delta>0 else -1
        self._merge_rotate(direction)

    def _merge_rotate(self, direction):
        if self.merge_state!="prep": return
        step=math.radians(self.rotate_step.get())*(1 if direction>0 else -1)
        self.merge_angle+=step
        self._redraw()

    def _merge_key(self, ev):
        k=ev.keysym.lower()
        if k in ("q","e"):
            self._merge_rotate(+1 if k=="e" else -1)
        elif k=="bracketleft":
            self.rotate_step.set(clamp(self.rotate_step.get()-0.5,0.1,15.0)); self._log(f"Step: {self.rotate_step.get():.1f}°")
        elif k=="bracketright":
            self.rotate_step.set(clamp(self.rotate_step.get()+0.5,0.1,15.0)); self._log(f"Step: {self.rotate_step.get():.1f}°")
        elif k=="escape":
            self._merge_cancel()

    def _merge_cancel(self):
        if self.merge_state!="prep": return
        self.merge_state=None
        self.merge_base=None; self.merge_move=None
        self.merge_init_vertices=None
        self.merge_best=(None,None)
        self.merge_dx=0.0; self.merge_dy=0.0; self.merge_angle=0.0
        self._update_wizard_bindings()
        self._log("Alignment cancelled.")
        self._redraw()

    def _mask_union_merge(self, Av, Bv):
        xs=[p[0] for p in Av+Bv]; ys=[p[1] for p in Av+Bv]
        xmin=min(xs); xmax=max(xs); ymin=min(ys); ymax=max(ys)
        pad=6.0; s=4.0
        rx0=int(math.floor(xmin-pad)); ry0=int(math.floor(ymin-pad))
        rw=int(math.ceil((xmax-xmin)+2*pad)); rh=int(math.ceil((ymax-ymin)+2*pad))
        if rw<=0 or rh<=0: return None
        W=int(rw*s); H=int(rh*s)
        if W<=4 or H<=4: return None
        def map_pts(pts):
            out=[]
            for x,y in pts:
                xx=(x-rx0)*s; yy=(y-ry0)*s
                out.append((int(round(xx)), int(round(yy))))
            return np.array(out, np.int32)
        Amsk=np.zeros((H,W), np.uint8)
        Bmsk=np.zeros((H,W), np.uint8)
        cv2.fillPoly(Amsk,[map_pts(Av)],255)
        cv2.fillPoly(Bmsk,[map_pts(Bv)],255)
        U=cv2.bitwise_or(Amsk,Bmsk)
        k=cv2.getStructuringElement(cv2.MORPH_ELLIPSE,(3,3))
        U=cv2.morphologyEx(U, cv2.MORPH_CLOSE, k, iterations=1)
        U=cv2.morphologyEx(U, cv2.MORPH_OPEN,  k, iterations=1)
        cnts,_=cv2.findContours(U, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not cnts: return None
        big=max(cnts, key=cv2.contourArea)
        peri=cv2.arcLength(big, True)
        eps=max(0.5*s, 0.002*peri)
        approx=cv2.approxPolyDP(big, eps, True).reshape(-1,2)
        merged=[]
        for x,y in approx:
            X=x/s+rx0; Y=y/s+ry0
            merged.append((float(X), float(Y)))
        merged=[(int(round(x)),int(round(y))) for (x,y) in merged]
        merged=simplify_chain(merged, eps=0.5)
        return merged

    def _merge_commit(self):
        if self.merge_state!="prep": return
        cand=self._merge_current_candidate()
        if cand is None: return
        ang, off, eg = self._merge_metrics(cand)
        ok_ang = ang <= float(self.merge_tol_angle.get())
        ok_off = off <= float(self.merge_tol_offset.get())
        ok_eg  = eg  <= float(self.merge_tol_endgap.get())
        if not (ok_ang and ok_off and ok_eg):
            if not messagebox.askyesno("Merge","Alignment outside tolerances. Proceed with mask-union merge?"):
                return
        A=self.merge_base; B=self.merge_move
        Av=list(A.polygon["vertices_px"]); Bv=list(cand)
        if signed_area(Av)<0: Av=list(reversed(Av))
        if signed_area(Bv)<0: Bv=list(reversed(Bv))
        merged=self._mask_union_merge(Av,Bv)
        if not merged or len(merged)<3:
            messagebox.showerror("Merge","Failed to compute union shape."); return
        H,W=self.image.shape[:2]
        cand_m=polygon_mask(merged, (H,W))
        tmp_sel=self.selected_area
        self.selected_area=A
        if not self._check_constraints(merged, ignore=B):
            self.selected_area=tmp_sel
            messagebox.showwarning("Merge","Merged polygon violates constraints. Adjust alignment or relax constraints.")
            return
        self.selected_area=tmp_sel
        A.polygon["vertices_px"]=[(int(round(x)),int(round(y))) for (x,y) in merged]
        A.polygon["is_circ"]=False
        A.polygon["circ_center_px"]=(None,None)
        A.polygon["circ_radius_px"]=None
        A.polygon_mask=cand_m
        self._update_metrics_and_inclusion(A, log=True)
        B.polygon=None; B.polygon_mask=None
        self._log(f"Merged '{A.name}' with '{B.name}'.")
        self._merge_cancel()
        self._redraw()

    def _compose_overlay(self, include_current=False, draw_current_center=False, draw_current_polygon=False):
        img=self.image.copy() if self.image is not None else None
        if img is None: return None
        base_h = mean_hue(COLORS[self.sel_color.get()]) if (self.input_type=="image" and hasattr(self,'sel_color')) else 105
        pal=build_palette(base_h)
        if self.contrast.get() and self.mask is not None:
            inv=cv2.bitwise_not(self.mask)
            img[inv>0]=(img[inv>0]*0.78).astype(np.uint8)
            overlay=img.copy()
            fill=np.zeros_like(img); fill[:]=pal["area_fill"]
            mask3=cv2.merge([self.mask,self.mask,self.mask])
            img=np.where(mask3>0,(0.74*overlay + 0.26*fill).astype(np.uint8), overlay)
            cnts,_=cv2.findContours(self.mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cv2.drawContours(img, cnts, -1, pal["area_edge"], 3)
        for idx,a in enumerate(self.areas, start=1):
            palA=build_palette(a.base_h if a.base_h is not None else (25*idx)%180)
            ov=img.copy()
            cv2.drawContours(ov,[a.contour.astype(np.int32)],-1,palA["area_fill"],cv2.FILLED)
            img=cv2.addWeighted(ov,0.18,img,0.82,0)
            cv2.drawContours(img,[a.contour.astype(np.int32)],-1,palA["area_edge"], 3 if a is self.selected_area else 2)
            if a.polygon:
                col_line=(0,255,0) if (a is self.selected_area and self.wizard_on.get() and self.merge_state!="prep") else palA["poly_edge"]
                col_vert=(0,255,0) if (a is self.selected_area and self.wizard_on.get() and self.merge_state!="prep") else palA["poly_vert"]
                if not a.included: col_line=(150,150,150); col_vert=(120,120,120)
                vp=a.polygon["vertices_px"]
                for i in range(len(vp)):
                    j=(i+1)%len(vp)
                    cv2.line(img,(int(vp[i][0]),int(vp[i][1])),(int(vp[j][0]),int(vp[j][1])),col_line,2)
                    cv2.circle(img,(int(vp[i][0]),int(vp[i][1])),3,col_vert,-1)
                x0,y0=a.contour[0]
                cv2.putText(img,a.name,(int(x0),int(y0)),cv2.FONT_HERSHEY_SIMPLEX,0.6,palA["poly_edge"],2,cv2.LINE_AA)
        if include_current and (self.curr_contour is not None or draw_current_polygon):
            palC=pal
            if self.curr_contour is not None:
                cnt=self.curr_contour.astype(np.int32).reshape(-1,2)
                ov=img.copy(); cv2.drawContours(ov,[cnt],-1,palC["area_fill"],cv2.FILLED)
                img=cv2.addWeighted(ov,0.22,img,0.78,0)
                cv2.drawContours(img,[cnt],-1,palC["area_edge"],2)
            if draw_current_polygon:
                pts=self.draw_pts[:]
                for i in range(len(pts)-1):
                    cv2.line(img,(int(pts[i][0]),int(pts[i][1])),(int(pts[i+1][0]),int(pts[i+1][1])),palC["poly_edge"],2)
                for (x,y) in pts:
                    cv2.circle(img,(int(x),int(y)),3,palC["poly_vert"],-1)
                if self.cursor_pt and pts:
                    cv2.line(img,(int(pts[-1][0]),int(pts[-1][1])),(int(self.cursor_pt[0]),int(self.cursor_pt[1])),palC["poly_edge"],1)
                    cv2.line(img,(int(pts[0][0]),int(pts[0][1])),(int(self.cursor_pt[0]),int(self.cursor_pt[1])),palC["poly_vert"],1)
        if draw_current_center and self.sel_center and self.radius:
            cv2.circle(img, self.sel_center, int(self.radius), pal["inner_circle"], 2)
        if self.wizard_on.get() and self.drag_preview is not None and self.selected_area is not None and self.merge_state!="prep":
            col=(0,255,0) if self.drag_valid else (0,0,255)
            pts=self.drag_preview
            for i in range(len(pts)):
                j=(i+1)%len(pts)
                cv2.line(img,(int(pts[i][0]),int(pts[i][1])),(int(pts[j][0]),int(pts[j][1])),col,2)
            for (x,y) in pts:
                cv2.circle(img,(int(x),int(y)),3,col,-1)
        if self.merge_state=="prep" and self.merge_init_vertices is not None and self.merge_base is not None and self.merge_move is not None:
            cand=self._merge_current_candidate()
            ia,ib=self.merge_best
            Av=list(self.merge_base.polygon["vertices_px"])
            if signed_area(Av)<0: Av=list(reversed(Av))
            ax1,ay1=Av[ia]; ax2,ay2=Av[(ia+1)%len(Av)]
            colA=pal["ghost_edge"]
            cv2.line(img,(int(ax1),int(ay1)),(int(ax2),int(ay2)),colA,3)
            if cand:
                ang,off,eg=self._merge_metrics(cand)
                ok_ang=ang <= float(self.merge_tol_angle.get())
                ok_off=off <= float(self.merge_tol_offset.get())
                ok_eg =eg  <= float(self.merge_tol_endgap.get())
                col=(0,255,0) if (ok_ang and ok_off and ok_eg) else (0,0,255)
                for i in range(len(cand)):
                    j=(i+1)%len(cand)
                    cv2.line(img,(int(cand[i][0]),int(cand[i][1])),(int(cand[j][0]),int(cand[j][1])),pal["ghost_poly"],2)
                bx1,by1=cand[ib]; bx2,by2=cand[(ib+1)%len(cand)]
                cv2.line(img,(int(bx1),int(by1)),(int(bx2),int(by2)),col,3)
                s=f"Δangle={ang:.2f}°, offset={off:.2f}px, endgap={eg:.2f}px"
                cv2.putText(img,s,(10,24),cv2.FONT_HERSHEY_SIMPLEX,0.7,pal['text'],2,cv2.LINE_AA)
                cv2.circle(img,(int(self.merge_pivot[0]),int(self.merge_pivot[1])),4,(255,255,255),-1)
        return img

    def _redraw(self, include_current=False, draw_current_center=False, draw_current_polygon=False):
        if self.image is None:
            self.canvas.delete("all"); return
        img=self._compose_overlay(include_current, draw_current_center, draw_current_polygon)
        if img is None: return
        self._show_image(img)

    def _show_image(self, bgr):
        self.canvas.delete("all")
        rgb=cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        H,W=rgb.shape[:2]
        disp_w=max(1,int(round(W*self.zoom))); disp_h=max(1,int(round(H*self.zoom)))
        rgb2=cv2.resize(rgb,(disp_w,disp_h),interpolation=cv2.INTER_NEAREST)
        self.tk_img=ImageTk.PhotoImage(Image.fromarray(rgb2))
        self.canvas.create_image(0,0,anchor="nw",image=self.tk_img)
        self.canvas.config(scrollregion=(0,0,disp_w,disp_h))

    def _event_xy_img(self, ev):
        xc=self.canvas.canvasx(ev.x); yc=self.canvas.canvasy(ev.y)
        return (int(xc/self.zoom), int(yc/self.zoom))
    def _pan_start(self, ev): self.canvas.scan_mark(ev.x, ev.y)
    def _pan_drag(self, ev):  self.canvas.scan_dragto(ev.x, ev.y, gain=1)
    def _zoom_generic(self, direction, ev):
        old=self.zoom; fac=1.1 if direction>0 else (1/1.1)
        new=clamp(old*fac, self.min_zoom, self.max_zoom)
        if abs(new-old)<1e-4: return
        mx=self.canvas.canvasx(ev.x)/old; my=self.canvas.canvasy(ev.y)/old
        self.zoom=new; self._redraw()
        disp_w=self.tk_img.width(); disp_h=self.tk_img.height()
        left=mx*new - ev.x; top=my*new - ev.y
        self.canvas.xview_moveto(clamp(left/max(1,disp_w),0.0,1.0))
        self.canvas.yview_moveto(clamp(top /max(1,disp_h),0.0,1.0))
    def _zoom_wheel(self, ev):
        direction=+1 if ev.delta>0 else -1
        self._zoom_generic(direction, ev)

if __name__=="__main__":
    app=SmartPolygonApp()
    app._run_step()
    app.mainloop()
